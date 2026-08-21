"""Normalise the judges' RDTII 2.1 Databases into our own output schema.

The two Database workbooks ARE the answer key, but their shape is nothing like our
submission CSV: one spreadsheet row can hold several Acts in a single cell, a parallel
list of timeframes, references spread across a variable number of unnamed columns, and
a free-prose justification that carries the provision citation inside the sentence.

This tool re-expresses that as a reference dataset with EXACTLY our export columns
(imported from the program, never re-typed here) so a run can be diffed against it
field-by-field. It is a read-only consumer of the workbooks and of our schema — it does
not touch the pipeline or the submission format.

Mapping decisions, and what is deliberately left blank:

    Economy            <- worksheet name (already the UN name our export writes)
    Law Name           <- "Act and/or practice", split on ';' (the panel's own list separator)
    Law Number / Ref     (blank — the panel has no such field; numbers sit inside the name)
    Last Amended       <- the "last amended …" clause of the aligned Timeframe segment
    Indicator ID       <- "6.1" -> "P6-I1"
    Article / Section  <- citations the panel names in its own justification prose
    Discovery Tag        (blank — the panel records no provenance)
    Location Reference   (blank — no page/paragraph anchors in the workbooks)
    Verbatim Snippet     (blank — the panel wrote commentary, not quoted provision text;
                          filling this from prose would fake a verbatim quote)
    Mapping Rationale  <- "Impact or comments on Acts or practices", verbatim
    Source URL         <- References block, aligned to the law when the counts allow
    Confidence           (blank — not a concept in the workbooks)
    Notes              <- the panel's Timeframe + Note cells, verbatim and labelled
    Pillar             <- Pillar_ID
    RDTII_Raw_Score    <- "Raw Score", verbatim
    Coverage           <- "Coverage", verbatim

Nothing is inferred: no score is derived, no law name is completed, no URL is guessed.

Usage:  python tools/build_reference_dataset.py
"""
from __future__ import annotations

import csv
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from backend.config import ROOT                                    # noqa: E402
from backend.export.csv_export import MASTER_EXTRA_COLUMNS         # noqa: E402
from backend.schemas import SUBMISSION_COLUMNS                     # noqa: E402

# Worksheet -> the workbook it lives in. Sheet names are already the names our export
# writes in the Economy column, so no translation table is needed.
SOURCES: dict[str, list[str]] = {
    "ESCAP-RDTII-2.1_ Round 1 Database.xlsx": ["Australia", "Malaysia", "Singapore"],
    "ESCAP-RDTII-2.1_ Round 2 Database.xlsx": ["China", "India", "Mongolia"],
}
OUT_CSV = ROOT / "data" / "ground_truth" / "rdtii_reference_p67.csv"

_IND_RE = re.compile(r"^([67])\.(\d+)$")
_URL_RE = re.compile(r"https?://[^\s;,\"'\]<>]+")

# Provision citations as the panel writes them, in its own prose. Chinese/Indian/Mongolian
# sheets say "Article 27" / "Article 20.1.5"; SG/AU/MY say "Section 199" / "Sections 25 and
# 26" / "Section 11(3)". Structural units too coarse to be a retrieval target (Chapter,
# Part, Division) are deliberately not matched — nor is bare "s 77" without a period, which
# collides with ordinary prose.
_NUM = r"\d{1,4}[A-Za-z]{0,3}(?:\.\d{1,3}){0,4}(?:\([^)\s]{1,10}\))*"
_CIT_RE = re.compile(
    r"\b(?P<kind>articles?|sections?|ss?\.|rules?|regulations?|clauses?|paragraphs?|paras?"
    r"|schedules?)\s*(?P<num>" + _NUM + r")"
    r"(?:\s*(?:and|,|&)\s*(?P<num2>" + _NUM + r"))?",
    re.I)
_APP_RE = re.compile(r"\b(?:australian\s+privacy\s+principle|APP)\s*(\d+(?:\.\d+)?)", re.I)
_BARE_PAREN_RE = re.compile(r"\((\d{2,3}[A-Z]{1,2})\)")           # AU drafting: "(187C)"

# Canonical singular label per citation kind, so "Sections 25 and 26" and "section 25"
# normalise to the same string without changing the panel's vocabulary.
_KIND_LABEL = {"article": "Article", "section": "Section", "s": "Section", "ss": "Section",
               "rule": "Rule", "regulation": "Regulation", "clause": "Clause",
               "paragraph": "Paragraph", "para": "Paragraph", "schedule": "Schedule"}

_AMENDED_RE = re.compile(r"last\s+(?:amend|amene)\w*\s*(?P<when>.+)$", re.I)
# "last amended in 2021" / "… on 1 February 2021" / "… in2024" (the panel's own typo)
_AMEND_FILLER_RE = re.compile(r"^(?:in|on|since|as\s+of|from)\s*", re.I)
_YEAR_RE = re.compile(r"^(?:1[89]|20)\d{2}$")   # "Regulations 2021" is a title, not a provision

# Tokens that carry no identity when matching a law name inside prose. Deliberately short:
# "amendment", "code", "practice", "standard" and "bill" DO distinguish instruments the
# panel lists side by side ("Personal Data Protection Act" vs "… (Amendment) Bill" vs
# "… Code of Practice"), so they must stay in play.
_STOP = {"the", "of", "on", "and", "or", "a", "an", "to", "for", "in", "act", "law", "laws",
         "no", "republic", "people", "peoples"}


@dataclass
class RefRow:
    economy: str
    pillar: str
    indicator_id: str
    raw_score: str
    coverage: str
    law_name: str
    last_amended: str
    article_section: str
    rationale: str
    source_url: str
    notes: str
    extra_urls: list[str] = field(default_factory=list)


def _clean(cell) -> str:
    """Cell text with the panel's \xa0 placeholders treated as empty, trailing space gone."""
    if cell is None:
        return ""
    s = str(cell).replace("\xa0", " ")
    return "" if not s.strip() else "\n".join(ln.rstrip() for ln in s.strip().splitlines())


def _flat(cell) -> str:
    return re.sub(r"\s+", " ", _clean(cell)).strip()


def _segments(cell) -> list[str]:
    """The panel's ';'-separated parallel lists (laws, timeframes)."""
    return [s for s in (re.sub(r"\s+", " ", p).strip(" ;") for p in _clean(cell).split(";")) if s]


def _indicator_id(raw: str) -> str | None:
    m = _IND_RE.match(raw.strip())
    return f"P{m.group(1)}-I{m.group(2)}" if m else None


def _score(cell) -> str:
    s = _flat(cell)
    try:
        return f"{float(s):g}"
    except ValueError:
        return s


def _last_amended(timeframe: str) -> str:
    """The amendment clause of a Timeframe like 'Since December 1967 last amended in 2021'.

    Blank when the panel records no amendment — "Since 2021" alone does not assert that a
    law was never amended, so it must not become "Original" here.
    """
    m = _AMENDED_RE.search(timeframe)
    if not m:
        return ""
    return _AMEND_FILLER_RE.sub("", m.group("when").strip()).strip(" .,;")


@dataclass
class Citation:
    start: int          # where the citation begins in the prose
    end: int            # where it ends — the law it belongs to is usually named just after
    label: str          # "Section 199", "Article 20.1.5", "APP 8"


def _citations(text: str) -> list[Citation]:
    """Every provision the panel cites in its own prose, in order, deduplicated.

    A bare year is rejected: "Telecommunications Regulations 2021" is a law title, and
    reading it as "Regulation 2021" would invent a provision that does not exist.
    """
    found: list[Citation] = []
    for m in _CIT_RE.finditer(text):
        label = _KIND_LABEL.get(m.group("kind").lower().rstrip("s.").rstrip("s") or "section",
                                "Section")
        for g in ("num", "num2"):
            num = m.group(g)
            if num and not _YEAR_RE.match(num):
                found.append(Citation(m.start(), m.end(), f"{label} {num}"))
    for m in _APP_RE.finditer(text):
        found.append(Citation(m.start(), m.end(), f"APP {m.group(1)}"))
    for m in _BARE_PAREN_RE.finditer(text):
        found.append(Citation(m.start(), m.end(), f"Section {m.group(1)}"))
    seen, out = set(), []
    for c in sorted(found, key=lambda c: c.start):
        if c.label.lower() not in seen:
            seen.add(c.label.lower())
            out.append(c)
    return out


def _law_tokens(name: str) -> set[str]:
    return {w for w in re.findall(r"[a-z]{4,}", name.lower()) if w not in _STOP}


# The panel ties a citation to its instrument with "of the <title>" — "Article 18.6 of the
# Law on Transparency of Public Information", "Section 145(1) of the Companies Act".
_TIE_RE = re.compile(r"^[\s,.]*(?:\([^)]{0,60}\)[\s,]*)?of\s+(?:the\s+|this\s+)?", re.I)


def _acronym(law: str) -> str:
    """PDPA from 'Personal Data Protection Act 2012' — the panel cites laws this way
    ('Section 25 of the PDPA'), and a title-token match alone would miss those."""
    words = [w for w in re.findall(r"[A-Za-z]+", law) if w.lower() not in
             {"of", "the", "on", "and", "for", "in", "to", "a", "an", "no"}]
    return "".join(w[0].upper() for w in words) if len(words) >= 2 else ""


def _title_score(law: str, window: str, acronym: str) -> int:
    """How strongly a stretch of prose names this law.

    Title matching has to be loose: the panel paraphrases ("Law on Personal Data
    Protection" is cited as "Law on the Protection of Personal Information"), so an exact
    substring test would attribute almost nothing. An acronym hit counts as a full match.
    """
    if acronym and re.search(rf"\b{re.escape(acronym)}\b", window):
        return 99
    tokens = _law_tokens(law)
    if len(tokens) < 2:
        return 0
    hit = len(tokens & {w for w in re.findall(r"[a-z]{4,}", window.lower())})
    # Half the title, but never more than 4 words: Mongolia's "Attachment to the Order of
    # the Minister … on Regulation on Technical Requirements for Processing Sensitive,
    # Biometric, and Genetic Data" has 13 distinctive tokens, and demanding 7 of them makes
    # it unmatchable while a 3-word title next to it matches loosely everywhere.
    return hit if hit >= max(2, min((len(tokens) + 1) // 2, 4)) else 0


def _mention_positions(laws: list[str], acronyms: list[str], prose: str) -> list[list[int]]:
    """Where each law is named in the prose, by sliding a ~14-word window over it.

    Scoring is competitive: a window counts as naming a law only if that law is the single
    best match there. Without this, an Act and its Amendment Act — whose distinctive tokens
    can be identical — both "match" every mention of either, and the citation lands on
    whichever happens to sort last.
    """
    words = [(m.start(), m.group(0)) for m in re.finditer(r"[A-Za-z]{2,}", prose)]
    raw: list[list[int]] = [[] for _ in laws]
    for i in range(len(words)):
        window = " ".join(w for _, w in words[i:i + 14])
        scores = [_title_score(law, window, acr) for law, acr in zip(laws, acronyms)]
        top = max(scores)
        if top and scores.count(top) == 1:
            raw[scores.index(top)].append(words[i][0])
    return [[p for j, p in enumerate(hits) if j == 0 or p - hits[j - 1] > 40] for hits in raw]


def _sections_per_law(laws: list[str], prose: str) -> list[str]:
    """Attribute each cited provision to the law it belongs to.

    One law in the row -> every citation belongs to it; that is what the row asserts.

    Several laws -> two rules, in order:
      1. an explicit tie right after the citation ("Article 3 … of the Regulation on
         Technical Requirements …") decides it;
      2. otherwise the law named most recently BEFORE the citation, because the panel
         writes a paragraph that introduces an instrument and then cites its sections
         inline ("Under the PDPA, … (section 26) …").
    Searching backwards only is what keeps Mongolia 6.3 straight — that row discusses two
    instruments in one paragraph, and a nearest-either-direction match pulls the second
    law's articles onto the first. A citation with neither signal is left unattributed
    rather than assigned to the row's first law.
    """
    cites = _citations(prose)
    if not cites or not laws:
        return [""] * len(laws)
    if len(laws) == 1:
        return ["; ".join(c.label for c in cites)]
    acronyms = [_acronym(law) for law in laws]
    if len(set(a for a in acronyms if a)) != len([a for a in acronyms if a]):
        acronyms = [""] * len(laws)          # two laws share an acronym: too ambiguous to use
    mentions = _mention_positions(laws, acronyms, prose)
    buckets: list[list[str]] = [[] for _ in laws]
    for c in cites:
        tie = _TIE_RE.match(prose[c.end:c.end + 160])
        target = None
        if tie:
            window = prose[c.end + tie.end():c.end + tie.end() + 120]
            scores = [_title_score(law, window, acr) for law, acr in zip(laws, acronyms)]
            if max(scores) and scores.count(max(scores)) == 1:
                target = scores.index(max(scores))
        if target is None:
            before = [(p, i) for i, ps in enumerate(mentions)
                      for p in [max([q for q in ps if q < c.start], default=None)]
                      if p is not None]
            if before:
                target = max(before)[1]
            else:
                # The paragraph sometimes cites first and names its instrument after
                # ("According to Section 26, an organisation may only transfer … under the
                # PDPA"), so fall back to the nearest mention ahead of the citation.
                after = [(p, i) for i, ps in enumerate(mentions)
                         for p in [min([q for q in ps if q > c.start], default=None)]
                         if p is not None]
                if after:
                    target = min(after)[1]
        if target is not None:
            buckets[target].append(c.label)
    return ["; ".join(b) for b in buckets]


def _notes(timeframe: str, note: str) -> str:
    """The panel's own Timeframe/Note cells, labelled so they cannot be mistaken for ours.

    Timeframe is preserved whole here even though its amendment clause is also parsed into
    Last Amended — the full string ("Since December 1967 last amended in 2021") is the
    panel's wording and nothing should have to be reconstructed from the derived field.
    """
    parts = []
    if timeframe:
        parts.append(f"BGK Timeframe: {timeframe}")
    if note:
        parts.append(f"BGK Note: {re.sub(r'\n{3,}', '\n\n', note)}")
    return " | ".join(parts)


def _refs_bounds(header: tuple) -> tuple[int, int]:
    """References occupy columns 7 .. (the 'Note' column) - 1. That boundary moves per
    worksheet (Australia 12, Malaysia 11, India 12 …), so it is located, never assumed."""
    for i, h in enumerate(header):
        if h and "note" in str(h).lower():
            return 7, i
    return 7, len(header)


def parse_sheet(ws, economy: str) -> list[RefRow]:
    rows_in = list(ws.iter_rows(values_only=True))
    ref_lo, note_i = _refs_bounds(rows_in[0])
    out: list[RefRow] = []
    for r in rows_in[1:]:
        pillar = _flat(r[0])
        indicator = _indicator_id(_flat(r[1]))
        if pillar not in ("6", "7") or not indicator:
            continue                                  # pillar banner row, or another pillar
        laws = _segments(r[3])
        timeframes = _segments(r[6])
        prose = _clean(r[5])
        urls: list[str] = []
        for cell in r[ref_lo:note_i]:
            urls += [u.rstrip(".,;") for u in _URL_RE.findall(_clean(cell))]
        sections = _sections_per_law(laws, prose)
        # The panel keeps laws and timeframes as parallel lists and they align on every
        # pillar-6/7 row in both workbooks; references align on most. Where they do not,
        # the whole reference set is carried on each row instead of pinning one URL to one
        # law — the surplus entries are commentary/mirrors, not the next law's source.
        aligned = len(urls) == len(laws)
        for i, law in enumerate(laws or [""]):
            out.append(RefRow(
                economy=economy, pillar=pillar, indicator_id=indicator,
                raw_score=_score(r[2]), coverage=_flat(r[4]), law_name=law,
                last_amended=_last_amended(timeframes[i]) if i < len(timeframes) else "",
                article_section=sections[i] if i < len(sections) else "",
                rationale=prose,
                source_url=urls[i] if aligned else "; ".join(urls),
                notes=_notes(timeframes[i] if i < len(timeframes) else "",
                             _clean(r[note_i]) if note_i < len(r) else ""),
            ))
    return out


def build() -> list[RefRow]:
    import openpyxl
    rows: list[RefRow] = []
    for filename, sheets in SOURCES.items():
        path = ROOT / filename
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for sheet in sheets:
            if sheet not in wb.sheetnames:
                raise SystemExit(f"{filename}: no worksheet {sheet!r}")
            rows += parse_sheet(wb[sheet], sheet)
        wb.close()
    return rows


def write_csv(rows: list[RefRow], path: Path = OUT_CSV) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=SUBMISSION_COLUMNS + MASTER_EXTRA_COLUMNS,
                           quoting=csv.QUOTE_ALL)
        w.writeheader()
        for r in rows:
            w.writerow({
                "Economy": r.economy,
                "Law Name": r.law_name,
                "Law Number / Ref": "",
                "Last Amended": r.last_amended,
                "Indicator ID": r.indicator_id,
                "Article / Section": r.article_section,
                "Discovery Tag": "",
                "Location Reference": "",
                "Verbatim Snippet": "",
                "Mapping Rationale": r.rationale,
                "Source URL": r.source_url,
                "Confidence": "",
                "Notes": r.notes,
                "Pillar": r.pillar,
                "RDTII_Raw_Score": r.raw_score,
                "Coverage": r.coverage,
            })
    return path


def report(rows: list[RefRow]) -> str:
    from collections import Counter
    econ = Counter(r.economy for r in rows)
    lines = [f"rows: {len(rows)}",
             f"laws: {len({(r.economy, r.law_name) for r in rows})}",
             f"with Article/Section: {sum(1 for r in rows if r.article_section)}",
             f"with Last Amended: {sum(1 for r in rows if r.last_amended)}",
             f"with Source URL: {sum(1 for r in rows if r.source_url)}"]
    for e in sorted(econ):
        ind = sorted({r.indicator_id for r in rows if r.economy == e})
        lines.append(f"  {e:<10} rows={econ[e]:>3}  indicators={len(ind)}  {' '.join(ind)}")
    return "\n".join(lines)


if __name__ == "__main__":
    rows = build()
    print(report(rows))
    print("written:", write_csv(rows))

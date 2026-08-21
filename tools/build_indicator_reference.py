"""Assemble one indicator reference for all twelve RDTII 2.1 pillars, from the four sources
the panel actually gave us.

Until now `backend/rdtii/indicators.py` carried nine hand-written indicators for pillars 6 and
7, and everything else was a blank. That was the right call while the round was pillars 6 and 7
only. It is the wrong call now: the final round accepts any pillar, and we were treating the
gap as a research problem when in fact the material has been in the repository all along —
spread across four files, none of which is sufficient on its own.

    ESCAP-RDTII-2.1_ Round 1 Database.xlsx, "RDTII 2.1 Methodology" sheet
        61 indicators with the SCORING CRITERIA and the possible scores. This is the operative
        test — "1) Local storage requirement for all sectors or personal data … 2) applied to a
        specific sector … 3) No requirement", scored 1 / 0.5 / 0.

    OUTPUT_TEMPLATE_FINAL_ROUND.xlsx, "Indicator Reference" sheet
        The same 61 codes with the indicator TITLE, the WEIGHT inside its pillar, and the
        per-indicator exceptions. Also the five mapping traps at the foot of the sheet.

    ESCAP-RDTII-2.1-guide.pdf, chapter 3 (pages 26-104)
        A page or more of NARRATIVE per indicator: what the measure is, why it is a barrier,
        and worked country examples. This is what turns a one-line criterion into something a
        grader can apply to an unfamiliar statute.

    ESCAP-RDTII-2.1- Non-regulatory indicators.pdf
        The 14 indicators derived from external databases and treaty membership, where the
        document says outright that "an automated data retrieval method is not required".
        These are OUT OF SCOPE for the tool, and knowing which they are is the difference
        between 47 indicators of real work and 61 with 14 unwinnable.

Output: data/rdtii/indicator_reference.json

    python tools/build_indicator_reference.py
    python tools/build_indicator_reference.py --show 6.2
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from backend.config import ROOT                                    # noqa: E402

ROUND1_DB = ROOT / "ESCAP-RDTII-2.1_ Round 1 Database.xlsx"
FINAL_TEMPLATE = ROOT / "Finalist Orientation" / "OUTPUT_TEMPLATE_FINAL_ROUND.xlsx"
GUIDE_DIR = ROOT / "RDTII 2.1 framework documentations"
GUIDE_PDF = GUIDE_DIR / "ESCAP-RDTII-2.1-guide.pdf"
NONREG_PDF = GUIDE_DIR / "ESCAP-RDTII-2.1- Non-regulatory indicators.pdf"
OUT_JSON = ROOT / "data" / "rdtii" / "indicator_reference.json"

_CODE = re.compile(r"^\d{1,2}\.\d{1,2}(?:\.\d)?$")


def _clean(v) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v).replace("\xa0", " ")).strip()


# ── 1 · scoring criteria, from the Methodology sheet ─────────────────────────────────
def methodology() -> dict[str, dict]:
    import openpyxl
    wb = openpyxl.load_workbook(ROUND1_DB, read_only=True, data_only=True)
    out: dict[str, dict] = {}
    pillar = ""
    for row in wb["RDTII 2.1 Methodology"].iter_rows(values_only=True):
        c = [_clean(x) for x in row] + [""] * 6
        if c[0] and not _CODE.match(c[1]):
            pillar = c[0]
        if not _CODE.match(c[1]):
            continue
        out[c[1]] = {"pillar": int(c[1].split(".")[0]), "pillar_name": pillar,
                     "category": c[2], "scoring_criteria": c[3], "possible_scores": c[4]}
    wb.close()
    return out


# ── 2 · titles, weights, exceptions and the traps, from the final-round template ──────
def template_reference() -> tuple[dict[str, dict], list[str]]:
    import openpyxl
    if not FINAL_TEMPLATE.exists():
        return {}, []
    wb = openpyxl.load_workbook(FINAL_TEMPLATE, read_only=True, data_only=True)
    out: dict[str, dict] = {}
    traps: list[str] = []
    in_traps = False
    for row in wb["Indicator Reference"].iter_rows(values_only=True):
        c = [_clean(x) for x in row] + [""] * 5
        if "mapping traps" in c[0].lower():
            in_traps = True
            continue
        if in_traps:
            if c[0]:
                traps.append(c[0])
            continue
        if not _CODE.match(c[1]):
            continue
        title, _, exception = c[2].partition("Exception:")
        out[c[1]] = {"title": title.strip(), "weight_in_pillar": c[3],
                     "exception": (exception or c[4]).strip()}
    wb.close()
    return out, traps


# ── 3 · which indicators the tool is not expected to answer ──────────────────────────
def non_regulatory() -> dict[str, str]:
    import pypdf
    if not NONREG_PDF.exists():
        return {}
    text = "\n".join((p.extract_text() or "") for p in pypdf.PdfReader(NONREG_PDF).pages)
    text = re.sub(r"[ \t]+", " ", text)
    out: dict[str, str] = {}
    # Each row starts on its own line with the code and then wraps across several lines:
    #   "6.5 Not in an agreement with binding commitments on data transfer.
    #    Data from checking the participation status of the binding agreements."
    # So the rows are cut at the next code rather than matched one at a time.
    starts = [(m.start(), m.group(1)) for m in re.finditer(r"\n\s*(\d{1,2}\.\d{1,2})\s", text)]
    for i, (pos, code) in enumerate(starts):
        end = starts[i + 1][0] if i + 1 < len(starts) else len(text)
        out[code] = re.sub(r"\s+", " ", text[pos:end]).strip()[:300]
    return out


# ── 4 · the narrative definition, from chapter 3 of the guide ────────────────────────
def guide_sections(titles: dict[str, str]) -> dict[str, str]:
    """Chapter 3 discusses each indicator under its TITLE, not its code, so the codes are
    attached by matching the title text. A title that never appears as a heading simply gets no
    narrative — recorded as absent rather than guessed at."""
    import pypdf
    if not GUIDE_PDF.exists():
        return {}
    reader = pypdf.PdfReader(GUIDE_PDF)
    pages = [(i + 1, reader.pages[i].extract_text() or "") for i in range(len(reader.pages))]
    body = "\n".join(t for i, t in pages if 26 <= i <= 104)
    # strip the running header that repeats on every page and would break heading detection
    body = re.sub(r"Chapter 3 ♦ RDTII 2\.1 pillars", " ", body)
    body = re.sub(r"\d{0,3}\s*Regional Digital Trade Integration Index 2\.1: A Guide\s*\d{0,3}", " ", body)
    flat = re.sub(r"[ \t]+", " ", body)

    # A title also appears mid-paragraph — "local storage requirements are often found in
    # different laws" sits inside the discussion of data RETENTION — so matching the title alone
    # attaches the wrong text to the code, which is worse than attaching none. The guide's own
    # house style gives a reliable heading signal: a section opens with the title followed
    # immediately by "This indicator …". Headings are taken that way; a title that never appears
    # in that form is recorded as having no narrative rather than guessed at.
    heading: list[tuple[int, str]] = []
    loose: dict[str, int] = {}
    for code, title in titles.items():
        if len(title) < 8:
            continue
        pat = re.compile(re.escape(title[:60]).replace(r"\ ", r"\s+"), re.I)
        for m in pat.finditer(flat):
            after = flat[m.end():m.end() + 60]
            if re.match(r"\s*(?:This indicator|The indicator|This sub-?indicator)", after, re.I):
                heading.append((m.start(), code))
            else:
                loose.setdefault(code, m.start())

    heading.sort()
    out: dict[str, tuple[str, str]] = {}
    for idx, (pos, code) in enumerate(heading):
        end = heading[idx + 1][0] if idx + 1 < len(heading) else min(len(flat), pos + 6000)
        chunk = re.sub(r"\s+", " ", flat[pos:end]).strip()
        if len(chunk) > len(out.get(code, ("", ""))[0]):
            out[code] = (chunk[:6000], "heading")

    # Some sections open differently ("Monitoring requirements can be explicit or indirect…").
    # Those are still useful, but they are a weaker match, so they are labelled as such rather
    # than being presented alongside a verified one as if the two were equally reliable.
    boundaries = sorted(p for p, _ in heading)
    for code, pos in loose.items():
        if code in out:
            continue
        nxt = next((b for b in boundaries if b > pos), min(len(flat), pos + 4000))
        chunk = re.sub(r"\s+", " ", flat[pos:min(nxt, pos + 4000)]).strip()
        if len(chunk) > 200:
            out[code] = (chunk[:4000], "loose")
    return out


def build() -> dict:
    meth = methodology()
    tmpl, traps = template_reference()
    nonreg = non_regulatory()
    titles = {c: (tmpl.get(c, {}).get("title") or meth[c]["category"]) for c in meth}
    narrative = guide_sections(titles)

    # The Methodology sheet turns out to carry ONLY the regulatory indicators — 1.1, 2.4, 4.4,
    # 6.5, 9.2, 12.10 and the rest of the treaty-membership set are simply absent from it. That
    # is a useful accident: the sheet already IS the scope of our task. Assert it, so a future
    # edition that mixes them in does not quietly hand us fourteen unwinnable indicators.
    leaked = sorted(set(meth) & set(nonreg))
    if leaked:
        print(f"WARNING: non-regulatory codes present in the Methodology sheet: {leaked}")

    indicators = {}
    for code, m in sorted(meth.items(), key=lambda kv: [int(x) for x in kv[0].split(".")]):
        t = tmpl.get(code, {})
        indicators[code] = {
            "code": code,
            "pillar": m["pillar"],
            "pillar_name": m["pillar_name"] or t.get("pillar_name", ""),
            "title": titles[code],
            "weight_in_pillar": t.get("weight_in_pillar", ""),
            "scoring_criteria": m["scoring_criteria"],
            "possible_scores": m["possible_scores"],
            "exception": t.get("exception", ""),
            "guide_narrative": narrative.get(code, ("", ""))[0],
            # "heading" = matched the guide's own section opening, so it is certainly this
            # indicator. "loose" = the title appeared in prose; useful, but check before relying.
            "narrative_match": narrative.get(code, ("", ""))[1],
        }
    return {
        "source": "RDTII 2.1 — Methodology sheet, final-round Indicator Reference, "
                  "guide chapter 3, non-regulatory indicator list",
        # Every indicator in `indicators` is in scope for the tool. The Methodology sheet turns
        # out to list ONLY the regulatory ones, so the scope question answers itself: what is in
        # the sheet is our work, and the fourteen below are not. They are recorded anyway so
        # that a reader wondering why 6.5 is missing finds the reason rather than a gap.
        "out_of_scope_non_regulatory": nonreg,
        "mapping_traps": traps,
        "indicators": indicators,
    }


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--show", help="print one indicator and exit")
    args = ap.parse_args()

    data = build()
    inds = data["indicators"]

    if args.show:
        print(json.dumps(inds.get(args.show, {"error": "no such code"}),
                         indent=1, ensure_ascii=False))
        return 0

    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(json.dumps(data, indent=1, ensure_ascii=False), encoding="utf-8")

    with_narr = [c for c in inds if inds[c]["narrative_match"] == "heading"]
    loose_n = [c for c in inds if inds[c]["narrative_match"] == "loose"]
    print(f"{len(inds)} indicators IN SCOPE (the Methodology sheet lists only regulatory ones) · "
          f"{len(data['out_of_scope_non_regulatory'])} non-regulatory, out of scope")
    print(f"scoring criteria: {sum(1 for v in inds.values() if v['scoring_criteria'])}/{len(inds)}")
    print(f"guide narrative : {len(with_narr)}/{len(inds)} verified headings, "
          f"{len(loose_n)} loose matches (labelled, not trusted)")
    print(f"mapping traps   : {len(data['mapping_traps'])}\n")
    by_pillar: dict[int, list[str]] = {}
    for c, v in inds.items():
        by_pillar.setdefault(v["pillar"], []).append(c)
    for p in sorted(by_pillar):
        codes = by_pillar[p]
        v = sum(1 for c in codes if inds[c]["narrative_match"] == "heading")
        lo = sum(1 for c in codes if inds[c]["narrative_match"] == "loose")
        print(f"  pillar {p:>2}: {len(codes):>2} in scope · {v:>2} verified + {lo:>2} loose   "
              f"{' '.join(codes)}")
    print(f"\nwritten: {OUT_JSON}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

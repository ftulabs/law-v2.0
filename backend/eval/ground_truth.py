"""Labelled evaluation set derived from the judges' own RDTII 2.1 Databases.

The hackathon ships no labelled training data, but the Database workbooks ARE a label set:
for every (economy, indicator) the panel recorded the Act(s) they accepted as evidence, a
prose justification that almost always names the exact provision ("According to Section 199,
every company must retain accounting records for 5 years…"), and the Raw Score.

This module turns that into machine-checkable targets:

    LabelRow(economy, indicator_id, laws=[...], sections=[...], kind=provision|absence)

`kind` matters as much as the target itself. A row whose justification reads "Singapore does
not implement ban on data transfer" is NOT a retrieval target — it is a recorded ABSENCE, and
a pipeline that returns a provision there is producing a false positive, not a win. Scoring
against the Database without that distinction would reward exactly the wrong behaviour.

Nothing here is used by the pipeline at runtime — it exists so retrieval changes can be
measured instead of argued about.

TWO workbooks, ten economies. Round 1 labels SG/AU/MY; Round 2 labels the seven final-round
economies the panel scored — CN, IN, ID, LA, MN, RU, TH. Both sheets have the identical
column layout, so the same parser reads both; what differs is HOW the panel writes a citation.
Round 1 says "Section 199", Round 2 says "Article 27" / "Article 12.7" / "Article 20(2)", and
until `_ARTICLE_RE` existed every Round-2 row parsed to zero provision targets — not an error,
just a label set that silently measured nothing.

Timor-Leste is on the final-round country list and has NO sheet in either workbook, so it
cannot be labelled here and cannot be measured. That is a fact about the panel's data, not a
gap to paper over: an unlabelled economy must keep the conservative retrieval budget.
"""
from __future__ import annotations

import json
import re
from dataclasses import asdict, dataclass, field
from pathlib import Path

from ..config import ROOT

DATABASE_XLSX = "ESCAP-RDTII-2.1_ Round 1 Database.xlsx"
ROUND2_XLSX = "ESCAP-RDTII-2.1_ Round 2 Database.xlsx"
SHEET_ECONOMY = {"Singapore": "SG", "Australia": "AU", "Malaysia": "MY"}
SHEET_ECONOMY_R2 = {"China": "CN", "India": "IN", "Indonesia": "ID", "Lao PDR": "LA",
                    "Mongolia": "MN", "Russian Federation": "RU", "Thailand": "TH"}
# (workbook, sheet→economy). A workbook that is not present is skipped, not an error: the
# Round-2 database arrived later than the Round-1 one and a clone may hold only one of them.
DATABASES = ((DATABASE_XLSX, SHEET_ECONOMY), (ROUND2_XLSX, SHEET_ECONOMY_R2))
OUT_JSON = ROOT / "data" / "ground_truth" / "rdtii_p67_labels.json"

_IND_RE = re.compile(r"^([67])\.(\d+)$")

# Provision references as the panel writes them, in the prose justification.
#   "Section 199" / "section 39 and 40" / "Sections 25 and 26" / "s 77"
#   "Section 11(3)" / "Section 45(2)(a)(i)" / "Section 3.1.1" (codes of practice)
#   "APP 8" / "Australian Privacy Principle 8"
#   "(187C)" — AU drafting cites a bare section number in brackets
_SECTION_RE = re.compile(
    r"\bsections?\s+(\d{1,3}[A-Za-z]{0,2}(?:\([^)]{1,8}\))*(?:\.\d+){0,3})"
    r"(?:\s*(?:and|,|&)\s*(\d{1,3}[A-Za-z]{0,2}(?:\([^)]{1,8}\))*))?",
    re.I)
_APP_RE = re.compile(r"\b(?:australian\s+privacy\s+principle|APP)\s*(\d+(?:\.\d+)?)", re.I)
_BARE_PAREN_RE = re.compile(r"\((\d{2,3}[A-Z]{1,2})\)")          # "(187C)"

# Round-2 economies are civil-law drafted and the panel cites them by ARTICLE:
#   "Article 27" / "Articles 25 and 26" / "Art. 5"
#   "Article 12.7" / "Article 10.1.3" — RU and MN number clause-wise, which is exactly how
#                                       the extractor labels those provisions
#   "Article 20(2)" / "Article 7.1.3, part (d)"
# Deliberately NOT extended to "Chapter N", which several Round-2 justifications also cite:
# harness.section_key() returns None for structural headings, so a Chapter label could never
# match any provision and would only inflate prov_expected with unreachable targets — that
# depresses measured recall without any pipeline being at fault.
_ARTICLE_RE = re.compile(
    r"\b(?:articles?|art\.)\s*(\d{1,3}[A-Za-z]{0,2}(?:\.\d+){0,3}(?:\([^)]{1,8}\))*)"
    r"(?:\s*(?:and|,|&)\s*(\d{1,3}[A-Za-z]{0,2}(?:\.\d+){0,3}(?:\([^)]{1,8}\))*))?",
    re.I)

# Justifications that record an ABSENCE rather than a provision.
_ABSENCE_RE = re.compile(
    r"\b(?:does not (?:implement|have|impose|mandate|prescribe|expressly)"
    r"|do not (?:implement|have|impose)"
    r"|no\s+\w+(?:\s+\w+){0,3}\s+(?:requirements?|provisions?|measures?|laws?)?\s*"
    r"(?:were|was|are|is)?\s*(?:found|identified|specified)"
    r"|there (?:are|is) no\b"
    r"|not (?:mandated|required|specified)"
    r"|currently does not|no countries have been specified)",
    re.I)


@dataclass
class LabelRow:
    economy: str
    indicator_id: str          # P6-I1 …
    pillar: int
    raw_score: float | None
    coverage: str              # "Horizontal" | sector name — the panel's own scope note
    kind: str                  # "provision" (a citable target) | "absence" (correct answer = none)
    laws: list[str] = field(default_factory=list)        # law names the panel accepted
    sections: list[str] = field(default_factory=list)    # provision refs named in the justification
    portal_urls: list[str] = field(default_factory=list)  # official-portal references only
    other_urls: list[str] = field(default_factory=list)   # third-party mirrors / commentary
    impact: str = ""

    @property
    def key(self) -> str:
        return f"{self.economy}:{self.indicator_id}:{self.coverage or 'Horizontal'}"


def _indicator_id(raw: str) -> str | None:
    m = _IND_RE.match((raw or "").strip())
    return f"P{m.group(1)}-I{m.group(2)}" if m else None


def _split_laws(cell: str) -> list[str]:
    """The 'Act and/or practice' cell lists one or more instruments, separated by ';' or
    newlines, often with the newline INSIDE a name ('Personal Data Protection\n(Amendment)
    Bill'). Split on ';' first, then collapse internal whitespace."""
    out = []
    for part in re.split(r";|\n{2,}", cell or ""):
        name = re.sub(r"\s+", " ", part).strip(" ; ")
        if len(name) >= 4:
            out.append(name)
    return out


def _split_urls(cells) -> tuple[list[str], list[str]]:
    """Reference cells hold several URLs (newline/;-separated) mixed with free prose. Return
    (official-portal urls, everything else) — the panel frequently cites a third-party mirror
    (mohre.um.edu.my, cyrilla.org) for a law that IS on the official portal, so a URL match is
    not a usable target; the law NAME is."""
    portal_pat = re.compile(
        r"(sso\.agc\.gov\.sg|legislation\.gov\.au|lom\.agc\.gov\.my|pdp\.gov\.my"
        # Round-2 official portals and the ministry/regulator hosts the panel cites as primary
        r"|flk\.npc\.gov\.cn|(?:^|\.)gov\.cn|cac\.gov\.cn|indiacode\.nic\.in|meity\.gov\.in"
        r"|peraturan\.go\.id|peraturan\.bpk\.go\.id|jdihn\.go\.id"
        r"|laoofficialgazette\.gov\.la|legalinfo\.mn"
        r"|pravo\.gov\.ru|krisdika\.go\.th|ratchakitcha\.soc\.go\.th)", re.I)
    official, other = [], []
    for c in cells:
        for u in re.findall(r"https?://[^\s;,\"'\]]+", str(c or "")):
            u = u.rstrip(".,;")
            (official if portal_pat.search(u) else other).append(u)
    return official, other


def _sections(text: str) -> list[str]:
    """Provision refs named in the panel's justification, normalised to bare labels
    ('199', '11(3)', 'APP 8', '3.1.1')."""
    out: list[str] = []
    for rx in (_SECTION_RE, _ARTICLE_RE):
        for m in rx.finditer(text or ""):
            for g in m.groups():
                if g:
                    out.append(g.strip(" ."))
    for m in _APP_RE.finditer(text or ""):
        out.append(f"APP {m.group(1)}")
    for m in _BARE_PAREN_RE.finditer(text or ""):
        out.append(m.group(1))
    seen, uniq = set(), []
    for s in out:
        k = s.lower()
        if k not in seen:
            seen.add(k)
            uniq.append(s)
    return uniq


def _kind(impact: str, sections: list[str]) -> str:
    """A row is an ABSENCE target when the justification records that nothing exists. A named
    provision wins over the absence phrasing: several justifications say 'Malaysia does not
    have a strict localisation law … however Section 129 requires …', which IS a target."""
    if sections:
        return "provision"
    return "absence" if _ABSENCE_RE.search(impact or "") else "provision"


def _load_workbook_labels(path: Path, sheet_economy: dict[str, str]) -> list[LabelRow]:
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows: list[LabelRow] = []
    for sheet, econ in sheet_economy.items():
        if sheet not in wb.sheetnames:
            continue
        for r in wb[sheet].iter_rows(values_only=True):
            ind = _indicator_id(str(r[1] or ""))
            if not ind or str(r[0] or "").strip() not in ("6", "7"):
                continue
            impact = re.sub(r"\s+", " ", str(r[5] or "")).strip()
            secs = _sections(impact)
            official, other = _split_urls(r[7:12])
            try:
                score = float(r[2]) if r[2] not in (None, "", "\xa0") else None
            except (TypeError, ValueError):
                score = None
            rows.append(LabelRow(
                economy=econ, indicator_id=ind, pillar=int(ind[1]), raw_score=score,
                coverage=re.sub(r"\s+", " ", str(r[4] or "")).strip(),
                kind=_kind(impact, secs),
                laws=_split_laws(str(r[3] or "")), sections=secs,
                portal_urls=official, other_urls=other, impact=impact,
            ))
    wb.close()
    return rows


def load_labels(xlsx: str | Path | None = None) -> list[LabelRow]:
    """Parse the Database workbooks into label rows for pillars 6 and 7.

    `xlsx` reads ONE workbook (Round-1 sheet map) and is kept for callers that want to pin a
    single file; the default reads every workbook in DATABASES that is present on disk.
    """
    if xlsx is not None:
        return _load_workbook_labels(Path(xlsx), SHEET_ECONOMY)
    rows: list[LabelRow] = []
    for name, sheets in DATABASES:
        path = ROOT / name
        if path.exists():
            rows.extend(_load_workbook_labels(path, sheets))
    return rows


def labelled_economies() -> set[str]:
    """Economies the panel's own databases actually score — the only ones a retrieval budget
    can be MEASURED for. Everything else must keep the conservative default."""
    return {e for _, sheets in DATABASES for e in sheets.values()}


def export(path: Path | None = None) -> Path:
    out = Path(path) if path else OUT_JSON
    out.parent.mkdir(parents=True, exist_ok=True)
    rows = load_labels()
    out.write_text(json.dumps([asdict(r) for r in rows], indent=1, ensure_ascii=False),
                   encoding="utf-8")
    return out


def summary(rows: list[LabelRow]) -> dict:
    from collections import Counter
    by_econ = Counter(r.economy for r in rows)
    by_kind = Counter(f"{r.economy}/{r.kind}" for r in rows)
    laws = {(r.economy, law) for r in rows for law in r.laws}
    return {
        "rows": len(rows),
        "by_economy": dict(by_econ),
        "by_kind": dict(by_kind),
        "distinct_laws": len(laws),
        "provision_targets": sum(len(r.sections) for r in rows if r.kind == "provision"),
        "indicators_covered": len({(r.economy, r.indicator_id) for r in rows}),
    }


if __name__ == "__main__":
    rows = load_labels()
    print(json.dumps(summary(rows), indent=1))
    print("written:", export())
